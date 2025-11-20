#!/usr/bin/env python3.11
import openpyxl
import mysql.connector
from datetime import datetime
import os

DATABASE_URL = os.getenv('DATABASE_URL')

if not DATABASE_URL:
    print('❌ DATABASE_URL não encontrada')
    exit(1)

# Parse DATABASE_URL: mysql://user:pass@host:port/dbname
url_parts = DATABASE_URL.replace('mysql://', '').split('@')
user_pass = url_parts[0].split(':')
host_db = url_parts[1].split('/')
host_port = host_db[0].split(':')

db_config = {
    'user': user_pass[0],
    'password': user_pass[1],
    'host': host_port[0],
    'port': int(host_port[1]) if len(host_port) > 1 else 3306,
    'database': host_db[1].split('?')[0]
}

print('Conectando ao banco de dados...')
conn = mysql.connector.connect(**db_config)
cursor = conn.cursor()

print('Lendo arquivo Excel...')
wb = openpyxl.load_workbook('/home/ubuntu/upload/Mega-Sena.xlsx')
ws = wb.active

print(f'Total de linhas: {ws.max_row}')

draws_to_insert = []
skipped = 0

for row_idx in range(2, ws.max_row + 1):  # Skip header row
    try:
        row = list(ws[row_idx])
        
        contest = row[0].value
        date_str = row[1].value
        ball1 = row[2].value
        ball2 = row[3].value
        ball3 = row[4].value
        ball4 = row[5].value
        ball5 = row[6].value
        ball6 = row[7].value
        winners6 = row[8].value or 0
        
        # Parse date
        if isinstance(date_str, str):
            date_parts = date_str.split('/')
            draw_date = datetime(
                int(date_parts[2]),
                int(date_parts[1]),
                int(date_parts[0])
            )
        else:
            draw_date = date_str
        
        # Validate balls
        balls = [ball1, ball2, ball3, ball4, ball5, ball6]
        if any(b is None or b < 1 or b > 60 for b in balls):
            print(f'⚠️  Sorteio {contest}: bolas inválidas')
            skipped += 1
            continue
        
        # Check for duplicates
        if len(set(balls)) != 6:
            print(f'⚠️  Sorteio {contest}: bolas duplicadas')
            skipped += 1
            continue
        
        draws_to_insert.append((
            contest,
            draw_date,
            ball1, ball2, ball3, ball4, ball5, ball6,
            winners6,
            0  # prize
        ))
        
    except Exception as e:
        print(f'⚠️  Erro na linha {row_idx}: {e}')
        skipped += 1

print(f'Sorteios válidos: {len(draws_to_insert)}')
print(f'Sorteios ignorados: {skipped}')

if draws_to_insert:
    print('Inserindo dados no banco...')
    
    insert_query = """
    INSERT INTO draws (contest, drawDate, ball1, ball2, ball3, ball4, ball5, ball6, winners6, prize)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    # Insert in batches
    batch_size = 100
    for i in range(0, len(draws_to_insert), batch_size):
        batch = draws_to_insert[i:i+batch_size]
        cursor.executemany(insert_query, batch)
        conn.commit()
        print(f'Inseridos {min(i + batch_size, len(draws_to_insert))} de {len(draws_to_insert)}')
    
    print('✅ Importação concluída com sucesso!')

cursor.close()
conn.close()
